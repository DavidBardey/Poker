from datetime import datetime, timedelta
from pathlib import Path
from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
BASE_DIR = Path(__file__).resolve().parent
EXPORTS_DIR = BASE_DIR / "exports"
EXPORTS_DIR.mkdir(exist_ok=True)

GAME_STATE = {"configured": False, "players": [], "active_players": [], "eliminated_players": [], "events": [], "schedule": [], "detailed_schedule": [], "minutes_per_level": 15, "current_level": 1, "buy_in": 1000, "max_rebuys": 0, "end_time": "", "total_chips": 0}

def recommended_levels(nb_players:int,total_chips:int)->int:
    base=8
    if nb_players>=6: base+=2
    if nb_players>=9: base+=1
    if total_chips>=20000: base+=1
    if total_chips>=40000: base+=1
    if total_chips>=80000: base+=1
    return min(base,15)

def build_dynamic_blind_schedule(starting_stack:int,levels:int)->list[str]:
    base_sb=max(10,starting_stack//100)
    allowed_steps=[10,15,20,25,30,40,50,75,100,150,200,300,400,500,600,800,1000,1200,1500,2000,2500,3000,4000,5000,6000,8000,10000]
    def closest_step(x:float)->int: return min(allowed_steps,key=lambda v:abs(v-x))
    sbs=[]; current=base_sb
    for _ in range(levels):
        sbs.append(closest_step(current)); current*=1.5
    schedule=[]; used=set()
    for sb in sbs:
        text=f"{sb} / {sb*2}"
        if text not in used: schedule.append(text); used.add(text)
    while len(schedule)<levels:
        last_sb=int(schedule[-1].split('/')[0].strip())
        for mult in (1.5,2):
            next_sb=closest_step(last_sb*mult); text=f"{next_sb} / {next_sb*2}"
            if text not in used:
                schedule.append(text); used.add(text); break
        else: break
    return schedule[:levels]

def compute_blinds_timing(nb_players:int,end_time_text:str,buy_in:int,max_rebuys:int):
    now=datetime.now()
    try: target=datetime.strptime(end_time_text,'%H:%M')
    except ValueError: raise ValueError("L'heure limite doit être au format HH:MM, par exemple 23:30.")
    end_dt=now.replace(hour=target.hour,minute=target.minute,second=0,microsecond=0)
    if end_dt<=now: end_dt+=timedelta(days=1)
    total_minutes=int((end_dt-now).total_seconds()//60)
    if total_minutes<15: raise ValueError('Le temps restant est trop court pour proposer une structure de blinds.')
    total_chips=nb_players*buy_in*(1+max_rebuys)
    levels=recommended_levels(nb_players,total_chips)
    minutes_per_level=max(5,total_minutes//levels)
    schedule=build_dynamic_blind_schedule(buy_in,levels)
    detailed=[]; start=now
    for i, blind in enumerate(schedule,start=1):
        level_end=start+timedelta(minutes=minutes_per_level)
        detailed.append({'level':i,'blind':blind,'start':start.strftime('%H:%M'),'end':level_end.strftime('%H:%M')})
        start=level_end
    return {'now':now.strftime('%H:%M'),'end':end_dt.strftime('%H:%M'),'total_minutes':total_minutes,'levels':levels,'minutes_per_level':minutes_per_level,'schedule':schedule,'detailed_schedule':detailed,'buy_in':buy_in,'max_rebuys':max_rebuys,'total_chips':total_chips}

def record_event(action:str,killed_player:str='',killer:str='',recave:str=''):
    GAME_STATE['events'].append({'date':datetime.now().strftime('%Y-%m-%d'),'heure':datetime.now().strftime('%H:%M:%S'),'niveau':GAME_STATE.get('current_level',1),'action':action,'joueur_kille':killed_player,'killeur':killer,'recave':recave})

def current_winner():
    if len(GAME_STATE.get('active_players', [])) == 1:
        return GAME_STATE['active_players'][0]
    return ''

def recaves_used(player: str) -> int:
    return sum(
        1 for e in GAME_STATE.get('events', [])
        if e.get('action') == 'Kill'
        and e.get('joueur_kille') == player
        and e.get('recave') == 'Oui'
    )

def recaves_remaining(player: str) -> int:
    return max(0, int(GAME_STATE.get('max_rebuys', 0)) - recaves_used(player))

def build_ranking() -> dict:
    ranking = {}
    winner = current_winner()
    if winner:
        ranking[winner] = 1
        rank = 2
        for player in reversed(GAME_STATE.get('eliminated_players', [])):
            if player not in ranking:
                ranking[player] = rank
                rank += 1
    return ranking

def format_rank(rank):
    if not rank:
        return ''
    return '1er' if rank == 1 else f'{rank}e'

@app.route('/')
def index(): return render_template('index.html')

@app.route('/blinds/setup', methods=['GET','POST'])
def blinds_setup():
    if request.method=='POST':
        try:
            nb_players=int(request.form['nb_players']); end_time=request.form['end_time'].strip(); buy_in=int(request.form['buy_in']); max_rebuys=int(request.form['max_rebuys'])
            players=[]
            for i in range(1,nb_players+1):
                name=request.form.get(f'player_{i}','').strip()
                if not name: raise ValueError(f"Le joueur {i} n'a pas de nom.")
                players.append(name)
            if len(players)!=len(set(players)): raise ValueError('Deux joueurs ne peuvent pas avoir exactement le même nom.')
            result=compute_blinds_timing(nb_players,end_time,buy_in,max_rebuys)
            GAME_STATE.update({'configured':True,'players':players,'active_players':players.copy(),'eliminated_players':[],'events':[],'schedule':result['schedule'],'detailed_schedule':result['detailed_schedule'],'minutes_per_level':result['minutes_per_level'],'current_level':1,'buy_in':buy_in,'max_rebuys':max_rebuys,'end_time':result['end'],'total_chips':result['total_chips'],'total_minutes':result['total_minutes'],'levels':result['levels']})
            record_event('Début de partie')
            return render_template('blinds_summary.html', state=GAME_STATE)
        except Exception as exc:
            return render_template('blinds_setup.html', error=str(exc))
    return render_template('blinds_setup.html', error=None)

@app.route('/blinds/clock')
def blinds_clock():
    if not GAME_STATE['configured']: return render_template('index.html', error="Veuillez d'abord configurer la partie.")
    return render_template('blinds_clock.html', state=GAME_STATE)

@app.route('/api/level', methods=['POST'])
def api_level():
    data=request.get_json(force=True); direction=data.get('direction')
    if direction=='next': GAME_STATE['current_level']=min(GAME_STATE['current_level']+1,len(GAME_STATE['schedule'])); record_event(f"Passage au niveau {GAME_STATE['current_level']}")
    elif direction=='previous': GAME_STATE['current_level']=max(GAME_STATE['current_level']-1,1); record_event(f"Retour au niveau {GAME_STATE['current_level']}")
    return jsonify({'ok':True,'state':GAME_STATE})

@app.route('/api/recaves')
def api_recaves():
    return jsonify({
        'max_rebuys': GAME_STATE.get('max_rebuys', 0),
        'used': {p: recaves_used(p) for p in GAME_STATE.get('players', [])},
        'remaining': {p: recaves_remaining(p) for p in GAME_STATE.get('players', [])},
    })

@app.route('/api/kill', methods=['POST'])
def api_kill():
    data=request.get_json(force=True); killed_player=data.get('killed_player','').strip(); killer=data.get('killer','').strip(); outcome=data.get('outcome','').strip()
    if not killed_player: return jsonify({'ok':False,'error':'Il manque le joueur killé.'}),400
    if not killer: return jsonify({'ok':False,'error':'Il manque le joueur qui a fait le kill.'}),400
    if killed_player==killer: return jsonify({'ok':False,'error':'Le joueur killé ne peut pas être son propre killeur.'}),400
    if killed_player not in GAME_STATE['active_players']: return jsonify({'ok':False,'error':"Ce joueur n'est pas actif."}),400
    if killer not in GAME_STATE['active_players']: return jsonify({'ok':False,'error':'Le killeur doit être un joueur encore actif.'}),400
    if outcome not in ['rebuy','out']: return jsonify({'ok':False,'error':'Résultat invalide.'}),400
    if outcome == 'rebuy' and recaves_remaining(killed_player) <= 0:
        return jsonify({'ok': False, 'error': f'{killed_player} a déjà utilisé toutes ses recaves. Il doit être éliminé.'}), 400
    if outcome=='rebuy': record_event('Kill',killed_player=killed_player,killer=killer,recave='Oui')
    else:
        record_event('Kill',killed_player=killed_player,killer=killer,recave='Non')
        GAME_STATE['active_players'].remove(killed_player)
        if killed_player not in GAME_STATE['eliminated_players']: GAME_STATE['eliminated_players'].append(killed_player)
        record_event('Élimination définitive',killed_player=killed_player,killer=killer,recave='Non')
    return jsonify({'ok':True,'state':GAME_STATE,'winner':current_winner(), 'recaves_remaining': {p: recaves_remaining(p) for p in GAME_STATE.get('players', [])}})

@app.route('/api/winner')
def api_winner():
    return jsonify({'winner': current_winner(), 'active_players': GAME_STATE.get('active_players', []), 'ranking': build_ranking()})

@app.route('/export')
def export_excel():
    if not GAME_STATE['configured']: return 'Aucune partie configurée.',400
    winner=current_winner()
    wb=Workbook(); header_fill=PatternFill(start_color='C9A227',end_color='C9A227',fill_type='solid')
    ws=wb.active; ws.title='Journal'; ws.append(['Date','Heure','Niveau','Action','Joueur killé','Killé par','Recave'])
    for cell in ws[1]: cell.font=Font(bold=True); cell.fill=header_fill; cell.alignment=Alignment(horizontal='center')
    for e in GAME_STATE['events']: ws.append([e.get('date',''),e.get('heure',''),e.get('niveau',''),e.get('action',''),e.get('joueur_kille',''),e.get('killeur',''),e.get('recave','')])
    for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=max(len(str(c.value)) if c.value is not None else 0 for c in col)+3
    ws2=wb.create_sheet('Résumé joueurs'); ws2.append(['Joueur','Kills réalisés','Kills subis','Recaves','Statut final','Classement'])
    for cell in ws2[1]: cell.font=Font(bold=True); cell.fill=header_fill; cell.alignment=Alignment(horizontal='center')
    ranking = build_ranking()
    for player in GAME_STATE['players']:
        kills_realises=sum(1 for e in GAME_STATE['events'] if e.get('action')=='Kill' and e.get('killeur')==player)
        kills_subis=sum(1 for e in GAME_STATE['events'] if e.get('action')=='Kill' and e.get('joueur_kille')==player)
        recaves=recaves_used(player)
        statut='Vainqueur' if player==winner else ('Éliminé' if player in GAME_STATE['eliminated_players'] else 'Encore en jeu')
        classement=format_rank(ranking.get(player))
        ws2.append([player,kills_realises,kills_subis,recaves,statut,classement])
    for col in ws2.columns: ws2.column_dimensions[col[0].column_letter].width=max(len(str(c.value)) if c.value is not None else 0 for c in col)+3
    ws3=wb.create_sheet('Paramètres'); ws3.append(['Paramètre','Valeur']); ws3.append(['Date export',datetime.now().strftime('%Y-%m-%d %H:%M:%S')]); ws3.append(['Joueurs',', '.join(GAME_STATE['players'])]); ws3.append(['Vainqueur', winner or 'Non déterminé']); ws3.append(['Cave initiale',GAME_STATE['buy_in']]); ws3.append(['Recaves max',GAME_STATE['max_rebuys']]); ws3.append(['Heure de fin prévue',GAME_STATE['end_time']]); ws3.append(['Nombre de niveaux',len(GAME_STATE['schedule'])]); ws3.append(['Durée par niveau',f"{GAME_STATE['minutes_per_level']} minutes"]); ws3.append(['Jetons théoriques',GAME_STATE['total_chips']])
    for col in ws3.columns: ws3.column_dimensions[col[0].column_letter].width=max(len(str(c.value)) if c.value is not None else 0 for c in col)+3
    filename=f"resume_soiree_poker_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"; filepath=EXPORTS_DIR/filename; wb.save(filepath)
    return send_file(filepath, as_attachment=True, download_name=filename)

if __name__=='__main__': app.run(debug=True)
